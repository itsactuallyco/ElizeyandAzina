#!/usr/bin/env python3
"""
Story Garden — turn stories.xlsx into stories.json

    python3 build.py

Edit stories.xlsx, run this, commit both files. That is the whole workflow.
You can also skip the spreadsheet entirely and edit stories.json by hand;
just don't do both at once, or the next build will overwrite your hand edits.

Needs: pip install openpyxl
"""

import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is missing. Run:  pip install openpyxl")

HERE = Path(__file__).parent
XLSX = HERE / "stories.xlsx"
JSON = HERE / "stories.json"

# Values the site knows how to draw. Anything else falls back to a default
# and gets reported below, so a typo shows up here instead of on screen.
SYMBOLS = {"trade", "shield", "sword", "lantern", "torchwater", "numbers", "school",
           "hospital", "tools", "light", "medbook", "gears", "peacesword", "gold",
           "compass", "scales", "quran", "city", "cipher", "heartbook", "twobooks",
           "flask", "globe", "heart", "astrolabe", "quill", "observatory", "scroll",
           "ney", "dome", "oud", "map", "lawscroll", "gathering", "bow"}
HEADWEAR = {"turban", "hijab", "helmet", "crown", "sultan", "sikke", "cap"}
ICONS = {"brave", "leaders", "healers", "stars", "makers", "explorers", "thinkers", "gentle"}

DEFAULTS = dict(symbol="scroll", headwear="turban",
                robe="#CFE4DA", wc="#FDFAF5", tint="#CFE4DA")

warnings = []


def warn(msg):
    warnings.append(msg)


def clean(v):
    """Normalise smart quotes and stray whitespace; keep paragraph breaks."""
    if v is None:
        return ""
    s = str(v)
    for a, b in [("\u2019", "'"), ("\u2018", "'"), ("\u201c", '"'),
                 ("\u201d", '"'), ("\u2014", " — "), ("\u00a0", " ")]:
        s = s.replace(a, b)
    s = re.sub(r"[ \t]+", " ", s)
    return s.strip()


def rows(ws):
    """Yield each row as a dict keyed by the header text."""
    it = ws.iter_rows(values_only=True)
    header = [clean(h) for h in next(it)]
    for r in it:
        row = {header[i]: r[i] for i in range(len(header)) if i < len(r)}
        if any(clean(v) for v in row.values()):
            yield row


def colour(v, fallback, who, field):
    c = clean(v).upper()
    if re.fullmatch(r"#[0-9A-F]{6}", c):
        return c
    if c:
        warn(f"{who}: {field} '{c}' is not a #RRGGBB colour — using {fallback}")
    return fallback


def main():
    if not XLSX.exists():
        sys.exit(f"Can't find {XLSX.name}. Run this from inside the site folder.")

    wb = load_workbook(XLSX, data_only=True)
    for sheet in ("Stories", "Categories", "Pronunciation"):
        if sheet not in wb.sheetnames:
            sys.exit(f"{XLSX.name} has no '{sheet}' tab.")

    # ---- categories ----
    categories = []
    for r in rows(wb["Categories"]):
        cid = clean(r.get("Id"))
        if not cid:
            continue
        icon = clean(r.get("Icon")) or "thinkers"
        if icon not in ICONS:
            warn(f"Category '{cid}': icon '{icon}' is unknown — using thinkers")
            icon = "thinkers"
        categories.append(dict(
            id=cid,
            name=clean(r.get("Name")) or cid,
            blurb=clean(r.get("Blurb")),
            icon=icon,
            tint=colour(r.get("Colour"), "#CFE4DA", f"Category '{cid}'", "Colour"),
        ))
    if not categories:
        sys.exit("The Categories tab is empty.")
    valid_cats = {c["id"] for c in categories}

    # ---- people ----
    people = []
    seen = set()
    for r in rows(wb["Stories"]):
        name = clean(r.get("Name"))
        if not name:
            continue
        if name in seen:
            warn(f"{name}: appears more than once — only the first is used")
            continue
        seen.add(name)

        story = str(r.get("Children's Story") or "")
        paras = [clean(p) for p in story.split("\n") if clean(p)]
        if len(paras) < 2:
            warn(f"{name}: story needs at least two paragraphs "
                 f"(the last becomes 'Something to remember') — skipped")
            continue

        story_mode = None
        sm_text = str(r.get("Story Mode") or "")
        sm_paras = [clean(p) for p in sm_text.split("\n") if clean(p)]
        if sm_paras and len(sm_paras) < 2:
            warn(f"{name}: Story Mode needs at least two paragraphs — ignored for now")
        elif sm_paras:
            story_mode = dict(paras=sm_paras[:-1], lesson=sm_paras[-1])

        cat = clean(r.get("Category"))
        if cat not in valid_cats:
            fallback = categories[0]["id"]
            warn(f"{name}: category '{cat or '(blank)'}' is not on the Categories "
                 f"tab — filed under {fallback}")
            cat = fallback

        sym = clean(r.get("Symbol")) or DEFAULTS["symbol"]
        if sym not in SYMBOLS:
            warn(f"{name}: symbol '{sym}' is unknown — using {DEFAULTS['symbol']}")
            sym = DEFAULTS["symbol"]

        wear = clean(r.get("Headwear")) or DEFAULTS["headwear"]
        if wear not in HEADWEAR:
            warn(f"{name}: headwear '{wear}' is unknown — using {DEFAULTS['headwear']}")
            wear = DEFAULTS["headwear"]

        people.append(dict(
            name=name,
            area=clean(r.get("Area of Impact")),
            cat=cat,
            obj=sym,
            wear=wear,
            robe=colour(r.get("Robe Colour"), DEFAULTS["robe"], name, "Robe Colour"),
            wc=colour(r.get("Headwear Colour"), DEFAULTS["wc"], name, "Headwear Colour"),
            tint=colour(r.get("Card Colour"), DEFAULTS["tint"], name, "Card Colour"),
            paras=paras[:-1],
            lesson=paras[-1],
            storyMode=story_mode,
        ))

    if not people:
        sys.exit("No usable rows on the Stories tab.")

    # ---- pronunciation ----
    say = {}
    ipa = {}
    for r in rows(wb["Pronunciation"]):
        written, spoken = clean(r.get("Written")), clean(r.get("Say It Like"))
        if written and spoken:
            say[written] = spoken
        written_ipa = clean(r.get("IPA"))
        if written and written_ipa:
            ipa[written] = written_ipa

    # ---- write ----
    data = dict(categories=categories, people=people, pronunciations=say, ipa=ipa)
    JSON.write_text(json.dumps(data, ensure_ascii=False, indent=1), encoding="utf-8")

    empty = [c["name"] for c in categories
             if not any(p["cat"] == c["id"] for p in people)]
    if empty:
        warn("These categories have nobody in them and will show '0 stories': "
             + ", ".join(empty))

    with_story_mode = sum(1 for p in people if p["storyMode"])
    print(f"Wrote {JSON.name}")
    print(f"  {len(people)} people across {len(categories)} categories")
    print(f"  {with_story_mode} have Story Mode text")
    print(f"  {len(say)} pronunciation entries, {len(ipa)} with IPA")
    for c in categories:
        print(f"    {c['name']}: {sum(1 for p in people if p['cat'] == c['id'])}")

    if warnings:
        print(f"\n{len(warnings)} thing(s) to look at:")
        for w in warnings:
            print("  - " + w)
    else:
        print("\nNo problems found.")


if __name__ == "__main__":
    main()
